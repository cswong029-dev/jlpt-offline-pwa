from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
# JLPT Grammar.xlsx 位於 Japanese(AI) 專案根目錄（jlpt-offline-pwa 的上一層）
XLSX = ROOT.parent / "JLPT Grammar.xlsx"
OUT = ROOT / "data" / "packs" / "builtin-grammar" / "grammar.json"

LEVELS = ["N5", "N4", "N3", "N2", "N1"]


def norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip())


def token_from_title(title: str) -> str:
    # Mirrors gen_builtin_grammar_quizzes.py extraction logic.
    return re.split(r"[／・（(\s]", title)[0].strip()


def split_variants(title: str) -> list[str]:
    # For cases like "ちゃいけない・じゃいけない", create 2 variants.
    # Keep only non-empty fragments.
    parts = [p.strip() for p in re.split(r"[・]", title) if p and str(p).strip()]
    if not parts:
        return [token_from_title(title)]
    # If extraction found something else, still keep up to 2 variants.
    return parts[:2]


EN2ZH_PHRASES = [
    ("in order to", "為了"),
    ("as soon as", "一...就..."),
    ("must not", "不可以"),
    ("must", "必須"),
    ("should", "應該"),
    ("can", "可以"),
    ("only", "只有"),
    ("just", "只是"),
    ("because", "因為"),
    ("if", "如果"),
    ("when", "當...時"),
    ("while", "在...期間"),
    ("after", "在...之後"),
    ("before", "在...之前"),
    ("until", "直到"),
    ("even if", "即使"),
    ("even though", "雖然"),
    ("seem", "看起來"),
    ("look", "看起來"),
    ("probably", "大概"),
    ("I think", "我認為"),
    ("to be", "是"),
    ("and", "和"),
    ("or", "或"),
    ("not", "不"),
]


def en_to_zh_hint(text: str) -> str:
    s = norm(text)
    if not s:
        return ""
    out = s
    for en, zh in EN2ZH_PHRASES:
        out = re.sub(rf"\b{re.escape(en)}\b", zh, out, flags=re.IGNORECASE)
    out = out.replace("~", "")
    out = re.sub(r"\s*;\s*", "；", out)
    out = re.sub(r"\s*,\s*", "、", out)
    out = re.sub(r"\s+", " ", out).strip()
    return out


def example_sentence(token: str, variant_i: int) -> str:
    t = token.strip()

    # A few high-quality templates for common particle-like grammar points.
    # For the rest, we still generate a readable example sentence by quoting the grammar point.
    if t == "で":
        return "学校で日本語を勉強します。"
    if t == "が":
        return "わたしが先生です。"
    if t == "でも":
        return "雨でも行きます。"
    if t == "だけ":
        return "わたしは水だけ飲みます。"
    if t == "だ":
        return "これは学生だ。"
    if t == "です":
        return "これは学生です。"
    if t in ("でしょう", "だろう"):
        return "明日は雨でしょう。"

    # Generic fallback: keep it readable and ensure the token appears in the text
    # so that quiz cloze replacement can still work.
    # We add variant_i to slightly vary sentences.
    variants = [
        f"私は「{t}」の使い方を勉強します。",
        f"会話では「{t}」を自然に使います。",
        f"例文を見て「{t}」を覚えます。",
    ]
    return variants[variant_i % len(variants)]


def zh_summary(title: str, meaning_en: str) -> str:
    return f"用法重點：{title}。建議透過例句觀察語氣、接續與使用情境。"


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"找不到來源 Excel：{XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["full list"] if "full list" in wb.sheetnames else wb[wb.sheetnames[0]]

    counters = {lv: 0 for lv in LEVELS}
    items: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        level = norm(row[0])
        title = norm(row[2])
        reading = norm(row[3])
        meaning = norm(row[4])
        source = norm(row[10]) if len(row) > 10 else ""

        if level not in counters or not title:
            continue
        counters[level] += 1

        # id stable-ish by level order
        gid = f"bg-{level.lower()}-{counters[level]:03d}"
        token_main = token_from_title(title)
        variants = split_variants(title)
        # Ensure at least one variant is the token used by quiz generator.
        if token_main and token_main not in variants:
            variants = [token_main] + variants
        variants = variants[:2]

        examples = []
        for i in range(2):
            tk = variants[i] if i < len(variants) else token_main
            if not tk:
                continue
            ja = example_sentence(tk, i)
            examples.append({
                "ja": ja,
                "reading": reading,
                "zh": f"例句翻譯（參考）：{en_to_zh_hint(meaning) or '請依句意理解此文法功能'}",
            })

        summary = zh_summary(title, meaning)
        body = "使用提示：先看前後文，再判斷這個文法在句中的功能（敘述、推測、限制、條件、對比等）。"

        items.append(
            {
                "id": gid,
                "level": level,
                "title": title,
                "summary": summary,
                "body": body,
                "examples": examples,
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({"version": 1, "packId": "builtin-grammar", "items": items}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("written:", OUT)
    print("total:", len(items))
    print("by level:", counters)


if __name__ == "__main__":
    main()

