# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(r"g:\我的雲端硬碟\Japanese(AI)\jlpt-offline-pwa")
XLSX = Path(r"g:\我的雲端硬碟\Japanese(AI)\JLPT Grammar.xlsx")
OUT = ROOT / "data" / "packs" / "builtin-grammar" / "grammar.json"


def norm_text(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["full list"] if "full list" in wb.sheetnames else wb[wb.sheetnames[0]]

    items = []
    counters = {"N5": 0, "N4": 0, "N3": 0, "N2": 0, "N1": 0}

    for row in ws.iter_rows(values_only=True):
        level = norm_text(row[0])
        title = norm_text(row[2])
        reading = norm_text(row[3])
        meaning = norm_text(row[4])
        source = norm_text(row[10])

        if level not in counters or not title:
            continue

        counters[level] += 1
        gid = f"bg-{level.lower()}-{counters[level]:03d}"

        summary = meaning or f"{title} 的用法。"
        body_parts = []
        if reading:
            body_parts.append(f"romaji: {reading}")
        if source:
            body_parts.append(source)
        body = "\n".join(body_parts) if body_parts else ""

        example_ja = f"【{title}】"
        example_reading = reading if reading else title
        example_zh = meaning if meaning else f"{title} 的意思"

        items.append(
            {
                "id": gid,
                "level": level,
                "title": title,
                "summary": summary,
                "body": body,
                "examples": [
                    {
                        "ja": example_ja,
                        "reading": example_reading,
                        "zh": example_zh,
                    }
                ],
            }
        )

    data = {"version": 1, "packId": "builtin-grammar", "items": items}
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("written:", OUT)
    print("total:", len(items))
    print("by level:", counters)


if __name__ == "__main__":
    main()
